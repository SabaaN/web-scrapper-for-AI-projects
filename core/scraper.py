"""
core/scraper.py
===============
The scraper engine — all platforms, filtering, scoring, and output.
This is your original main.py refactored as a Python module.

Can be called two ways:
  1. From the CLI:       python main.py --query "RAG chatbot"
  2. From the chatbot:   from core.scraper import scrape
                         results = scrape("RAG chatbot")

The scrape() function is the single public entry point.
Everything else is internal.
"""

import json
import os
import time
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
import re

import httpx
from dotenv import load_dotenv
from selectolax.parser import HTMLParser

load_dotenv()

# ─── CONFIG ───────────────────────────────────────────────────────────────────

RESULTS_PER_KEYWORD = int(os.getenv("RESULTS_PER_KEYWORD", 20))
REQUEST_DELAY       = float(os.getenv("REQUEST_DELAY", 2.5))
REQUEST_TIMEOUT     = int(os.getenv("REQUEST_TIMEOUT", 15))
FREELANCER_LIMIT    = int(os.getenv("FREELANCER_LIMIT", 20))
REMOTIVE_API_URL    = os.getenv("REMOTIVE_API_URL", "https://remotive.com/api/remote-jobs")
RESULTS_DIR         = os.getenv("RESULTS_DIR", "results")
EXCEL_DIR           = os.getenv("EXCEL_DIR", "excel")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ─── KEYWORD MAP ──────────────────────────────────────────────────────────────

KEYWORD_MAP: dict[str, list[str]] = {
    "computer vision":  ["computer vision", "object detection", "YOLO", "image recognition"],
    "nlp":              ["NLP", "natural language processing", "text classification", "sentiment analysis", "deep learning", "AI model", "artificial intelligence"],
    "llm":              ["LLM", "ChatGPT API", "RAG pipeline", "fine tuning LLM", "deep learning", "AI model", "artificial intelligence"],
    "agent":            ["AI agent", "agentic AI", "autonomous agent"],
    "ml":               ["machine learning", "deep learning", "AI model", "artificial intelligence"],
    "mlops":            ["MLOps", "model deployment", "ML pipeline"],
    "ocr":              ["OCR", "document extraction AI", "document AI", "deep learning", "AI model", "artificial intelligence"],
    "recommendation":   ["recommendation system", "personalization engine", "deep learning", "AI model", "artificial intelligence"],
    "generative":       ["stable diffusion", "generative AI", "image generation AI", "deep learning", "AI model", "artificial intelligence"],
    "speech":           ["speech recognition", "text to speech AI", "voice AI", "deep learning", "AI model", "artificial intelligence"],
    "data science":     ["data science", "data analysis AI", "predictive analytics"],
}

PROJECT_TERMS: list[str] = [
    "project", "development", "developer", "development project",
    "implementation", "solution", "application", "system",
    "integration", "build", "building", "develop", "implement",
    "create", "prototype", "proof of concept", "POC", "MVP",
]

# ─── DATA MODEL ───────────────────────────────────────────────────────────────

@dataclass
class ProjectListing:
    title:            str
    platform:         str
    description:      str
    budget_min:       float | None
    budget_max:       float | None
    currency:         str | None
    skills:           list[str]
    bid_count:        int | None
    posted_date:      str | None
    url:              str
    source_type:      str
    keyword:          str = ""
    relevance_score:  float = 0.0
    preference_rank:  int | None = None

# ─── SCORING ──────────────────────────────────────────────────────────────────

PROJECT_INTENT_SIGNALS: dict[str, int] = {
    "project": 5, "development": 5, "developer": 4, "develop": 5,
    "developing": 5, "build": 5, "building": 5, "implement": 5,
    "implementation": 5, "integrate": 4, "integration": 4,
    "application": 3, "system": 3, "solution": 3, "prototype": 4,
    "proof of concept": 5, "poc": 5, "mvp": 5, "model": 3,
    "pipeline": 4, "api": 2, "engineer": 4,
}

AI_SIGNALS: dict[str, int] = {
    "artificial intelligence": 5, "machine learning": 5, "deep learning": 5,
    "computer vision": 5, "neural network": 4, "neural": 3, "opencv": 4,
    "tensorflow": 4, "pytorch": 4, "yolo": 4, "cnn": 3, "transformer": 3,
    "llm": 3, "large language model": 4, "rag": 3, "langchain": 3,
    "stable diffusion": 4, "generative ai": 4, "image generation": 4,
    "dataset": 3, "training": 3, "inference": 3, "fine-tuning": 3,
    "classification": 3, "segmentation": 3, "object detection": 4,
    "image recognition": 4, "facial recognition": 4, "anomaly detection": 4,
    "nlp": 3, "natural language": 3, "data science": 3, "predictive": 3,
    "ocr": 3, "recommendation system": 3, "ai model": 4, "ai-powered": 4,
    "ai-enabled": 4, "openai": 3, "gpt": 3, "huggingface": 3,
    "bot": 1, "automation": 1,
}

NEGATIVE_SIGNALS: dict[str, int] = {
    "photography": -4, "photo shoot": -4, "wedding photo": -4,
    "logo design": -5, "brand identity": -4, "t-shirt design": -4,
    "video editing": -3, "video production": -3, "social media marketing": -4,
    "seo": -4, "content writer": -4, "copywriter": -4, "blog writer": -4,
    "powerpoint": -4, "website design": -3, "web design": -3,
    "cold calling": -4, "sales representative": -4, "interior design": -4,
    "bookkeeping": -4, "typist": -3, "data entry": -2,
}

MIN_RELEVANCE_SCORE = 3

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def extract_query_concepts(query: str) -> tuple[list[str], list[str]]:
    query_lower = query.lower()
    matched_domains, related_terms = [], []
    for domain, terms in KEYWORD_MAP.items():
        if domain.lower() in query_lower:
            matched_domains.append(domain)
            related_terms.extend(terms)
            continue
        for term in terms:
            if term.lower() in query_lower:
                matched_domains.append(domain)
                related_terms.extend(terms)
                break
    related_terms   = list(dict.fromkeys(t.lower() for t in related_terms))
    matched_domains = list(dict.fromkeys(matched_domains))
    return matched_domains, related_terms


def extract_keywords(query: str) -> list[str]:
    query_lower = query.lower()
    matched: list[str] = []
    for category, keywords in KEYWORD_MAP.items():
        if category in query_lower or any(kw.lower() in query_lower for kw in keywords):
            matched.extend(keywords)
    seen: set[str] = set()
    unique = []
    for kw in matched:
        if kw not in seen:
            seen.add(kw)
            unique.append(kw)
    return unique if unique else [query]


def generate_project_queries(base_keywords: list[str]) -> list[str]:
    queries = []
    for keyword in base_keywords:
        queries.append(keyword)
        for term in PROJECT_TERMS:
            queries.append(f"{keyword} {term}")
    return list(dict.fromkeys(queries))


def deduplicate(listings: list[ProjectListing]) -> list[ProjectListing]:
    seen: set[str] = set()
    unique = []
    for listing in listings:
        if listing.url not in seen:
            seen.add(listing.url)
            unique.append(listing)
    return unique


def parse_budget_string(text: str) -> tuple[float | None, float | None, str | None]:
    if not text:
        return None, None, None
    symbol_map = {"£": "GBP", "$": "USD", "€": "EUR"}
    currency = None
    for sym, code in symbol_map.items():
        if sym in text:
            currency = code
            break
    numbers = re.findall(r"[\d,]+", text.replace(",", ""))
    nums = [float(n) for n in numbers if n]
    if not nums:
        return None, None, currency
    if len(nums) == 1:
        return nums[0], nums[0], currency
    return min(nums), max(nums), currency


def http_get(client: httpx.Client, url: str, **kwargs) -> httpx.Response | None:
    try:
        resp = client.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
        resp.raise_for_status()
        return resp
    except httpx.HTTPStatusError as e:
        print(f"    [HTTP {e.response.status_code}] {url}")
        return None
    except httpx.RequestError as e:
        print(f"    [Request error] {url} — {e}")
        return None

# ─── SCORING FUNCTIONS ────────────────────────────────────────────────────────

def score_relevance(listing: ProjectListing, query: str) -> float:
    title       = listing.title.lower()
    description = listing.description.lower()
    skills      = " ".join(listing.skills).lower()
    full_text   = f"{title} {description} {skills}"
    query_lower = query.lower()

    matched_domains, related_terms = extract_query_concepts(query)

    query_words = [
        w for w in query_lower.replace("-", " ").split()
        if len(w) > 2 and w not in {"the", "and", "for", "with", "projects", "project"}
    ]

    query_match = (sum(1 for w in query_words if w in full_text) / len(query_words) * 100) if query_words else 0

    ai_score = 0
    for term in related_terms:
        if term in full_text:
            ai_score += 15 if term in matched_domains else 8
    ai_match = min(ai_score, 100)

    project_score = sum(w for t, w in PROJECT_INTENT_SIGNALS.items() if t in full_text)
    project_match = min(project_score * 5, 100)

    skills_match = min(sum(20 for w in query_words if w in skills), 100)

    title_match = (sum(1 for w in query_words if w in title) / len(query_words) * 100) if query_words else 0

    negative_penalty = min(sum(abs(w) * 2 for t, w in NEGATIVE_SIGNALS.items() if t in full_text), 40)

    score = (
        query_match   * 0.25
        + ai_match    * 0.30
        + project_match * 0.15
        + skills_match  * 0.10
        + title_match   * 0.20
        - negative_penalty
    )
    return round(max(0, min(score, 100)), 2)


def score_listings(listings: list[ProjectListing], query: str) -> list[ProjectListing]:
    scored = []
    for listing in listings:
        listing.relevance_score = score_relevance(listing, query)
        if listing.relevance_score >= MIN_RELEVANCE_SCORE:
            scored.append(listing)
    dropped = len(listings) - len(scored)
    if dropped:
        print(f"  Filtered out {dropped} irrelevant listings (score < {MIN_RELEVANCE_SCORE})")
    return scored


def rank_preferences(listings: list[ProjectListing], top_n: int = 5) -> list[ProjectListing]:
    ranked = sorted(listings, key=lambda x: x.relevance_score, reverse=True)
    for listing in ranked:
        listing.preference_rank = None
    for rank, listing in enumerate(ranked[:top_n], start=1):
        listing.preference_rank = rank
    return ranked


def sort_listings(listings: list[ProjectListing]) -> list[ProjectListing]:
    return sorted(
        listings,
        key=lambda x: (
            -x.relevance_score,
            x.bid_count if x.bid_count is not None else 999,
            -(x.budget_max or 0),
        ),
    )

# ─── SCRAPERS ─────────────────────────────────────────────────────────────────

def scrape_freelancer(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    url = "https://www.freelancer.com/api/projects/0.1/projects/active/"
    params = {
        "query": keyword, "job_details": "true",
        "full_description": "true", "offset": 0,
        "limit": FREELANCER_LIMIT, "sort_field": "time_updated",
    }
    headers = {**HEADERS, "Accept": "application/json", "Freelancer-OAuth-V1": ""}
    resp = http_get(client, url, params=params, headers=headers)
    if not resp:
        return []
    try:
        data = resp.json()
    except Exception:
        return []
    projects = data.get("result", {}).get("projects", [])
    results: list[ProjectListing] = []
    for p in projects:
        budget    = p.get("budget", {}) or {}
        currency  = (p.get("currency") or {}).get("code")
        skills    = [j.get("name", "") for j in (p.get("jobs") or []) if j.get("name")]
        seo_url   = p.get("seo_url", "")
        bid_count = (p.get("bid_stats") or {}).get("bid_count")
        ts        = p.get("time_submitted")
        posted    = datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d") if ts else None
        results.append(ProjectListing(
            title       = p.get("title", "Untitled").strip(),
            platform    = "Freelancer.com",
            description = (p.get("description") or "")[:500].strip(),
            budget_min  = budget.get("minimum"),
            budget_max  = budget.get("maximum"),
            currency    = currency,
            skills      = skills,
            bid_count   = bid_count,
            posted_date = posted,
            url         = f"https://www.freelancer.com/projects/{seo_url}" if seo_url else "https://www.freelancer.com",
            source_type = "freelance",
            keyword     = keyword,
        ))
    return results


def scrape_remotive(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    AI_CATEGORIES = ["machine-learning", "data", "software-dev"]
    seen_ids: set[str] = set()
    results: list[ProjectListing] = []
    for category in AI_CATEGORIES:
        resp = http_get(client, REMOTIVE_API_URL, params={"search": keyword, "category": category}, headers=HEADERS)
        if not resp:
            continue
        try:
            data = resp.json()
        except Exception:
            continue
        for job in data.get("jobs", []):
            job_id = str(job.get("id", ""))
            if job_id in seen_ids:
                continue
            seen_ids.add(job_id)
            salary_str = job.get("salary", "") or ""
            budget_min, budget_max, currency = parse_budget_string(salary_str)
            tags = job.get("tags") or []
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(",")]
            desc_raw  = job.get("description", "") or ""
            desc_tree = HTMLParser(desc_raw)
            description = desc_tree.text(strip=True)[:500] if desc_raw else ""
            results.append(ProjectListing(
                title       = (job.get("title") or "").strip(),
                platform    = "Remotive",
                description = description,
                budget_min  = budget_min,
                budget_max  = budget_max,
                currency    = currency or "USD",
                skills      = tags,
                bid_count   = None,
                posted_date = job.get("publication_date"),
                url         = job.get("url") or "https://remotive.com",
                source_type = "remote_contract",
                keyword     = keyword,
            ))
    return results


def scrape_himalayas(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    resp = http_get(client, "https://himalayas.app/jobs/api/search", params={"q": keyword}, headers=HEADERS)
    if not resp:
        return []
    try:
        data = resp.json()
    except Exception:
        return []
    results: list[ProjectListing] = []
    for job in data.get("jobs", []):
        results.append(ProjectListing(
            title       = (job.get("title") or "").strip(),
            platform    = "Himalayas",
            description = (job.get("excerpt") or "")[:500].strip(),
            budget_min  = float(job["minSalary"]) if job.get("minSalary") else None,
            budget_max  = float(job["maxSalary"]) if job.get("maxSalary") else None,
            currency    = job.get("currency") or "USD",
            skills      = job.get("categories") or [],
            bid_count   = None,
            posted_date = job.get("pubDate"),
            url         = job.get("applicationLink") or "https://himalayas.app",
            source_type = "remote_contract",
            keyword     = keyword,
        ))
    return results


def scrape_remoteok(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    resp = http_get(client, "https://remoteok.com/api", headers=HEADERS)
    if not resp:
        return []
    try:
        data = resp.json()
    except Exception:
        return []
    jobs = data[1:] if isinstance(data, list) and len(data) > 1 else []
    keyword_lower = keyword.lower()
    results: list[ProjectListing] = []
    for job in jobs:
        title    = job.get("position", "").strip()
        tags     = job.get("tags") or []
        desc_raw = job.get("description", "") or ""
        if keyword_lower not in f"{title} {' '.join(tags)} {desc_raw}".lower():
            continue
        desc_tree = HTMLParser(desc_raw)
        results.append(ProjectListing(
            title       = title,
            platform    = "RemoteOK",
            description = desc_tree.text(strip=True)[:500] if desc_raw else "",
            budget_min  = float(job["salary_min"]) if job.get("salary_min") else None,
            budget_max  = float(job["salary_max"]) if job.get("salary_max") else None,
            currency    = "USD",
            skills      = tags,
            bid_count   = None,
            posted_date = job.get("date"),
            url         = job.get("url") or "https://remoteok.com",
            source_type = "remote_contract",
            keyword     = keyword,
        ))
    return results


def scrape_arbeitnow(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    resp = http_get(client, "https://www.arbeitnow.com/api/job-board-api", headers=HEADERS)
    if not resp:
        return []
    try:
        data = resp.json()
    except Exception:
        return []
    keyword_lower = keyword.lower()
    results: list[ProjectListing] = []
    for job in data.get("data", []):
        title    = (job.get("title") or "").strip()
        tags     = job.get("tags") or []
        desc_raw = job.get("description", "") or ""
        if keyword_lower not in f"{title} {' '.join(tags)} {desc_raw}".lower():
            continue
        desc_tree = HTMLParser(desc_raw)
        results.append(ProjectListing(
            title       = title,
            platform    = "Arbeitnow",
            description = desc_tree.text(strip=True)[:500] if desc_raw else "",
            budget_min  = None,
            budget_max  = None,
            currency    = "EUR",
            skills      = tags,
            bid_count   = None,
            posted_date = str(job.get("created_at")),
            url         = job.get("url") or "https://www.arbeitnow.com",
            source_type = "remote_contract",
            keyword     = keyword,
        ))
    return results

# ─── OUTPUT ───────────────────────────────────────────────────────────────────

def print_listing(i: int, p: ProjectListing) -> None:
    sep = "=" * 65
    print(f"\n{sep}")
    print(f"  #{i}  {p.title}")
    print(f"{sep}")
    print(f"  Platform:    {p.platform}  [{p.source_type}]")
    if p.budget_min and p.budget_max:
        curr = p.currency or ""
        budget_str = (
            f"{curr} {p.budget_min:,.0f}"
            if p.budget_min == p.budget_max
            else f"{curr} {p.budget_min:,.0f} – {p.budget_max:,.0f}"
        )
    else:
        budget_str = "Not specified"
    print(f"  Budget:      {budget_str}")
    if p.bid_count is not None:
        signal = "🟢 Low" if p.bid_count < 10 else ("🟡 Medium" if p.bid_count < 25 else "🔴 High")
        print(f"  Bids:        {p.bid_count}  {signal} competition")
    if p.skills:
        skills_str = ", ".join(p.skills[:7])
        if len(p.skills) > 7:
            skills_str += f" (+{len(p.skills) - 7} more)"
        print(f"  Skills:      {skills_str}")
    if p.description:
        desc = p.description[:280] + ("..." if len(p.description) > 280 else "")
        print(f"  Description: {desc}")
    print(f"  Score:       {p.relevance_score}")
    print(f"  Posted:      {p.posted_date or 'Unknown'}")
    print(f"  Link:        {p.url}")


def print_results(listings: list[ProjectListing]) -> None:
    if not listings:
        print("\n  No results found. Try a broader query.")
        return
    counts = Counter(p.platform for p in listings)
    print("\n" + "─" * 65)
    print(f"  Found {len(listings)} unique projects:")
    for platform, count in counts.most_common():
        print(f"    • {platform}: {count}")
    print("─" * 65)
    for i, listing in enumerate(listings, 1):
        print_listing(i, listing)
    print(f"\n{'─' * 65}")
    print(f"  Total: {len(listings)} | {datetime.now().strftime('%H:%M:%S')}")
    print("─" * 65)


ORDINALS = [
    "first", "second", "third", "fourth", "fifth", "sixth", "seventh",
    "eighth", "ninth", "tenth", "eleventh", "twelfth", "thirteenth",
    "fourteenth", "fifteenth", "sixteenth", "seventeenth", "eighteenth",
    "nineteenth", "twentieth",
]


def _next_filename(directory: str, ext: str) -> str:
    os.makedirs(directory, exist_ok=True)
    existing = [f for f in os.listdir(directory) if f.endswith(ext)]
    n = len(existing)
    stem = ORDINALS[n] if n < len(ORDINALS) else f"run_{n + 1}"
    return os.path.join(directory, f"{stem}{ext}")


def save_results(listings: list[ProjectListing], query: str) -> str:
    path = _next_filename(RESULTS_DIR, ".json")
    data = {
        "scraped_at":    datetime.utcnow().isoformat(),
        "query":         query,
        "total_results": len(listings),
        "results":       [asdict(p) for p in listings],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  ✓ JSON  → {path}")
    return path


def save_excel_results(listings: list[ProjectListing]) -> str | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [Excel skipped] pip install openpyxl")
        return None
    path = _next_filename(EXCEL_DIR, ".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Leads"
    headers = [
        "Title", "Platform", "Description", "Budget Min", "Budget Max",
        "Currency", "Skills", "Bid Count", "Posted Date", "URL",
        "Source Type", "Keyword", "Relevance Score", "Preference",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for p in listings:
        ws.append([
            p.title, p.platform, p.description, p.budget_min, p.budget_max,
            p.currency, ", ".join(p.skills), p.bid_count, p.posted_date,
            p.url, p.source_type, p.keyword, p.relevance_score, p.preference_rank,
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width  = min(max(max(len(str(c.value or "")) for c in col) + 2, 12), 60)
        ws.column_dimensions[letter].width = width
    wb.save(path)
    print(f"  ✓ Excel → {path}")
    return path

# ─── PUBLIC ENTRY POINT ───────────────────────────────────────────────────────

def scrape(
    query: str,
    save_json: bool  = True,
    save_excel: bool = True,
    silent: bool     = False,
) -> list[ProjectListing]:
    """
    The single public function the chatbot (and CLI) calls.

    Args:
        query:      The search query string
        save_json:  Whether to save results to results/ as JSON
        save_excel: Whether to save results to excel/ as XLSX
        silent:     If True, suppress terminal output (useful when called
                    from the chatbot — the UI handles display instead)

    Returns:
        List of scored, sorted ProjectListing objects
    """
    keywords = generate_project_queries(extract_keywords(query))

    if not silent:
        print(f"\n  Query:   \"{query}\"")
        print(f"  Sources: Freelancer.com · Remotive · Himalayas · RemoteOK · Arbeitnow\n")

    all_listings: list[ProjectListing] = []

    with httpx.Client(follow_redirects=True) as client:

        scrapers = [
            ("Freelancer.com", scrape_freelancer),
            ("Remotive",       scrape_remotive),
            ("Himalayas",      scrape_himalayas),
            ("RemoteOK",       scrape_remoteok),
            ("Arbeitnow",      scrape_arbeitnow),
        ]

        for name, fn in scrapers:
            if not silent:
                print(f"  [{name}] ...")
            for kw in keywords[:4]:
                results = fn(client, kw)
                if not silent:
                    print(f"    → '{kw}': {len(results)} projects")
                all_listings.extend(results)
                time.sleep(REQUEST_DELAY)

    unique = deduplicate(all_listings)
    if not silent:
        print(f"\n  Pre-filter: {len(unique)} unique")

    scored  = score_listings(unique, query)
    ranked  = rank_preferences(scored, top_n=5)
    sorted_ = sort_listings(ranked)

    if not silent:
        print(f"  Post-filter: {len(sorted_)} relevant\n")
        print_results(sorted_)

    if save_json:
        save_results(sorted_, query)
    if save_excel:
        save_excel_results(sorted_)

    return sorted_