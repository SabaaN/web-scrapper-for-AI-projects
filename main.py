"""
AI Project Leads Scraper
========================
Scrapes AI/ML project opportunities from:
  - Freelancer.com  (public JSON endpoint, no auth)
  - PeoplePerHour   (static HTML)
  - Toptal          (static HTML)
  - Remotive.io     (official free public API)

Usage:
  python main.py
  python main.py --query "computer vision object detection"
  python main.py --query "NLP chatbot" --save
"""

import argparse
import json
import os
import time
from dataclasses import asdict, dataclass
from datetime import datetime
import re

import httpx
from dotenv import load_dotenv
from selectolax.parser import HTMLParser

load_dotenv()

# ─── CONFIG ───────────────────────────────────────────────────────────────────

RESULTS_PER_KEYWORD = int(os.getenv("RESULTS_PER_KEYWORD", 20))
REQUEST_DELAY       = float(os.getenv("REQUEST_DELAY", 2.5))
REQUEST_TIMEOUT     = int(os.getenv("REQUEST_TIMEOUT", 15))
FREELANCER_LIMIT    = int(os.getenv("FREELANCER_LIMIT", 20))
PPH_BASE_URL        = os.getenv("PPH_BASE_URL", "https://www.peopleperhour.com")
TOPTAL_BASE_URL     = os.getenv("TOPTAL_BASE_URL", "https://www.toptal.com")
REMOTIVE_API_URL    = os.getenv("REMOTIVE_API_URL", "https://remotive.com/api/remote-jobs")
RESULTS_DIR         = os.getenv("RESULTS_DIR", "results")
EXCEL_DIR           = os.getenv("EXCEL_DIR", "excel")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ─── KEYWORD MAP ──────────────────────────────────────────────────────────────
# Maps broad user query terms to specific search keywords per platform.

KEYWORD_MAP: dict[str, list[str]] = {
    "computer vision":  ["computer vision", "object detection", "YOLO", "image recognition"],
    "nlp":              ["NLP", "natural language processing", "text classification", "sentiment analysis", "deep learning", "AI model", "artificial intelligence"],
    "llm":              ["LLM", "ChatGPT API", "RAG pipeline", "fine tuning LLM", "deep learning", "AI model", "artificial intelligence"],
    "agent":            ["AI agent", "agentic AI", "autonomous agent"],
    "ml":               ["machine learning", "deep learning", "AI model", "artificial intelligence", "deep learning", "AI model", "artificial intelligence"],
    "mlops":            ["MLOps", "model deployment", "ML pipeline"],
    "ocr":              ["OCR", "document extraction AI", "document AI", "deep learning", "AI model", "artificial intelligence"],
    "recommendation":   ["recommendation system", "personalization engine", "deep learning", "AI model", "artificial intelligence"],
    "generative":       ["stable diffusion", "generative AI", "image generation AI", "deep learning", "AI model", "artificial intelligence"],
    "speech":           ["speech recognition", "text to speech AI", "voice AI", "deep learning", "AI model", "artificial intelligence"],
    "data science":     ["data science", "data analysis AI", "predictive analytics"],
}
PROJECT_TERMS: list[str] = [
    "project",
    "development",
    "developer",
    "development project",
    "implementation",
    "solution",
    "application",
    "system",
    "integration",
    "build",
    "building",
    "develop",
    "implement",
    "create",
    "prototype",
    "proof of concept",
    "POC",
    "MVP",
]
# ─── DATA MODEL ───────────────────────────────────────────────────────────────

@dataclass
class ProjectListing:
    title:        str
    platform:     str
    description:  str
    budget_min:   float | None
    budget_max:   float | None
    currency:     str | None
    skills:       list[str]
    bid_count:    int | None    # Freelancer only — key competition signal
    posted_date:  str | None
    url:          str
    source_type:  str           # "freelance" | "remote_contract"
    keyword:      str = ""      # which keyword surfaced this result
    relevance_score:   float = 0.0
    preference_rank:   int | None = None

# ─── HELPERS ──────────────────────────────────────────────────────────────────
PROJECT_INTENT_SIGNALS: dict[str, int] = {
    "project": 5,
    "development": 5,
    "developer": 4,
    "develop": 5,
    "developing": 5,
    "build": 5,
    "building": 5,
    "implement": 5,
    "implementation": 5,
    "integrate": 4,
    "integration": 4,
    "application": 3,
    "system": 3,
    "solution": 3,
    "prototype": 4,
    "proof of concept": 5,
    "poc": 5,
    "mvp": 5,
    "model": 3,
    "pipeline": 4,
    "api": 2,
    "engineer": 4,
}

# ─── RELEVANCE FILTERING ──────────────────────────────────────────────────────

# Strong positive signals: real AI/ML work
AI_SIGNALS: dict[str, int] = {
    "artificial intelligence": 5, "machine learning": 5, "deep learning": 5,
    "computer vision": 5, "neural network": 4, "neural": 3, "opencv": 4,
    "tensorflow": 4, "pytorch": 4, "yolo": 4, "cnn": 3, "transformer": 3,
    "llm": 3, "large language model": 4, "rag": 3, "langchain": 3,
    "stable diffusion": 4, "generative ai": 4, "image generation": 4,
    "dataset": 3, "training": 3, "inference": 3, "fine-tuning": 3,
    "classification": 3, "segmentation": 3, "object detection": 4,
    "image recognition": 4, "facial recognition": 4, "anomaly detection": 4,
    "nlp": 3, "natural language": 3, "data science": 3, "predictive": 3,
    "ocr": 3, "recommendation system": 3, "ai model": 4, "ai-powered": 4,
    "ai-enabled": 4, "openai": 3, "gpt": 3, "huggingface": 3,
    "bot": 1, "automation": 1,           # weak alone, but helps in combination
}

# Strong negative signals: almost certainly not an AI/ML project
NEGATIVE_SIGNALS: dict[str, int] = {
    "photography": -4, "photo shoot": -4, "wedding photo": -4,
    "logo design": -5, "brand identity": -4, "t-shirt design": -4,
    "bus wrap": -4, "video editing": -3, "video production": -3,
    "video trim": -3, "social media marketing": -4, "seo": -4,
    "content writer": -4, "copywriter": -4, "blog writer": -4,
    "powerpoint": -4, "ppt": -4, "menu scraper": -3,
    "website design": -3, "web design": -3, "onlyfans": -6,
    "dating": -4, "adult": -6, "illustration": -3,
    "photo editing": -3, "photoshop editor": -4, "background removal": -2,
    "cold calling": -4, "sales representative": -4, "landscape architect": -4,
    "landscape design": -4, "interior design": -4, "3d rendering": -2,
    "architectural": -3, "chemical": -4, "chemist": -4,
    "tutoring": -2, "trainer needed": -3, "teacher": -2,
    "virtual assistant": -3, "administrative": -3, "bookkeeping": -4,
    "typist": -3, "data entry": -2, "manual testing": -2,
}

# Minimum score a listing must achieve to appear in results.
# Looking at the data: genuine AI projects score 15+, junk scores 0-14.
# "Automotive PR" → 0, "Classic Logo" → 0, "Video Editing" → 0
# "AI Voice Receptionist" → 39, "Chatbot NLP" → 39, "RAG platform" → 27
# Set at 15 — catches edge cases like "Offline Mobile Call Translator"
# (has speech recognition in skills, borderline but relevant to us)
MIN_RELEVANCE_SCORE = 3

def extract_query_concepts(query: str) -> tuple[list[str], list[str]]:
    """
    Identify the main AI domains in the user's query and return:

    1. matched_domains:
       The AI domains explicitly detected in the query.

    2. related_terms:
       The domain's related technical terms from KEYWORD_MAP.
    """

    query_lower = query.lower()

    matched_domains = []
    related_terms = []

    for domain, terms in KEYWORD_MAP.items():
        # Check the domain itself
        if domain.lower() in query_lower:
            matched_domains.append(domain)
            related_terms.extend(terms)

            continue

        # Also check whether one of the domain's terms
        # appears explicitly in the query.
        for term in terms:
            if term.lower() in query_lower:
                matched_domains.append(domain)
                related_terms.extend(terms)
                break

    # Remove duplicates while preserving order
    related_terms = list(dict.fromkeys(
        term.lower() for term in related_terms
    ))

    matched_domains = list(dict.fromkeys(matched_domains))

    return matched_domains, related_terms

def extract_keywords(query: str) -> list[str]:
    """
    Map a plain-English user query to a deduplicated list of search keywords.
    Falls back to the raw query if nothing in the map matches.
    """
    query_lower = query.lower()
    matched: list[str] = []

    for category, keywords in KEYWORD_MAP.items():
        if category in query_lower or any(kw.lower() in query_lower for kw in keywords):
            matched.extend(keywords)

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique = []
    for kw in matched:
        if kw not in seen:
            seen.add(kw)
            unique.append(kw)

    return unique if unique else [query]


def deduplicate(listings: list[ProjectListing]) -> list[ProjectListing]:
    """Remove duplicate listings by URL."""
    seen: set[str] = set()
    unique = []
    for listing in listings:
        if listing.url not in seen:
            seen.add(listing.url)
            unique.append(listing)
    return unique

def generate_project_queries(base_keywords: list[str]) -> list[str]:
    queries = []

    for keyword in base_keywords:
        queries.append(keyword)

        for project_term in PROJECT_TERMS:
            queries.append(f"{keyword} {project_term}")

    return list(dict.fromkeys(queries))

def score_relevance(listing: ProjectListing, query: str) -> float:
    """
    Calculate a 0-100 relevance score for a listing against the user's query.

    Components:
      - Query/title match
      - AI domain match
      - Project intent
      - Skills match
      - Negative/non-AI signals
    """

    title = listing.title.lower()
    description = listing.description.lower()
    skills = " ".join(listing.skills).lower()

    full_text = f"{title} {description} {skills}"
    query_lower = query.lower()

# Identify the user's requested AI domain
    matched_domains, related_terms = extract_query_concepts(query)

    query_words = [
        word for word in query_lower.replace("-", " ").split()
        if len(word) > 2
        and word not in {"the", "and", "for", "with", "projects", "project"}
    ]

    if query_words:
        matched_words = sum(1 for word in query_words if word in full_text)
        query_match = (matched_words / len(query_words)) * 100
    else:
        query_match = 0

# ── 2. AI/domain match ─────────────────────────────────────────

    ai_score = 0

    for term in related_terms:
        if term in full_text:
            # Explicitly requested domain gets more importance
            if term in matched_domains:
                ai_score += 15
            else:
                ai_score += 8

    ai_match = min(ai_score, 100)

    # ── 3. Project intent ────────────────────────────────────────────
    project_score = 0

    for term, weight in PROJECT_INTENT_SIGNALS.items():
        if term in full_text:
            project_score += weight

    project_match = min(project_score * 5, 100)

    # ── 4. Skills match ──────────────────────────────────────────────
    skills_match = 0

    for word in query_words:
        if word in skills:
            skills_match += 20

    skills_match = min(skills_match, 100)

    # ── 5. Title match ───────────────────────────────────────────────
    title_matches = sum(
        1 for word in query_words
        if word in title
    )

    if query_words:
        title_match = (title_matches / len(query_words)) * 100
    else:
        title_match = 0

    # ── 6. Negative signals ─────────────────────────────────────────
    negative_penalty = 0

    for term, weight in NEGATIVE_SIGNALS.items():
        if term in full_text:
            negative_penalty += abs(weight) * 2

    negative_penalty = min(negative_penalty, 40)

    # ── Final weighted score ─────────────────────────────────────────
    score = (
        query_match * 0.25
        + ai_match * 0.30
        + project_match * 0.15
        + skills_match * 0.10
        + title_match * 0.20
        - negative_penalty
    )

    return round(max(0, min(score, 100)), 2)

def rank_preferences(
    listings: list[ProjectListing],
    top_n: int = 5,
) -> list[ProjectListing]:
    """
    Rank the most relevant projects as user preferences.

    The top_n highest-scoring listings receive preference_rank
    values starting from 1. All remaining listings receive None.
    """

    ranked = sorted(
        listings,
        key=lambda listing: listing.relevance_score,
        reverse=True,
    )

    for listing in ranked:
        listing.preference_rank = None

    for rank, listing in enumerate(ranked[:top_n], start=1):
        listing.preference_rank = rank

    return ranked

def score_listings(
    listings: list[ProjectListing],
    query: str,
) -> list[ProjectListing]:
    """
    Score all listings and drop anything below MIN_RELEVANCE_SCORE.
    This is the single filtering gate — everything that passes has
    a genuine AI/ML signal.
    """
    scored = []
    for listing in listings:
        listing.relevance_score = score_relevance(listing, query)
        if listing.relevance_score >= MIN_RELEVANCE_SCORE:
            scored.append(listing)

    dropped = len(listings) - len(scored)
    if dropped:
        print(f"  Filtered out {dropped} irrelevant listings (score < {MIN_RELEVANCE_SCORE})")

    return scored

def sort_listings(listings: list[ProjectListing]) -> list[ProjectListing]:
    """
    Sort by:
      1. relevance_score descending (most relevant first)
      2. bid_count ascending (fewer bids = better opportunity)
      3. budget_max descending (higher budget first)
    """
    return sorted(
        listings,
        key=lambda x: (
            -x.relevance_score,
            x.bid_count if x.bid_count is not None else 999,
            -(x.budget_max or 0),
        ),
    )


def get(client: httpx.Client, url: str, **kwargs) -> httpx.Response | None:
    """
    Shared HTTP GET with timeout and basic error handling.
    Returns None on any failure so callers can skip gracefully.
    """
    try:
        resp = client.get(url, timeout=REQUEST_TIMEOUT, **kwargs)
        resp.raise_for_status()
        return resp
    except httpx.HTTPStatusError as e:
        print(f"    [HTTP {e.response.status_code}] {url}")
        return None
    except httpx.RequestError as e:
        print(f"    [Request error] {url} — {e}")
        return None

def parse_budget_string(text: str) -> tuple[float | None, float | None, str | None]:
    """
    Parse budget strings like '£50', '$100-$200', '€500+' into
    (min, max, currency_symbol).
    """
    import re
    if not text:
        return None, None, None

    symbol_map = {"£": "GBP", "$": "USD", "€": "EUR"}
    currency = None
    for sym, code in symbol_map.items():
        if sym in text:
            currency = code
            break

    # Extract all numbers from the string
    numbers = re.findall(r"[\d,]+", text.replace(",", ""))
    nums = [float(n) for n in numbers if n]

    if not nums:
        return None, None, currency
    if len(nums) == 1:
        return nums[0], nums[0], currency
    return min(nums), max(nums), currency


# ─── SCRAPER: FREELANCER.COM ──────────────────────────────────────────────────

def scrape_freelancer(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    """
    Calls Freelancer's public active-project search endpoint.
    No authentication or cookies required.
    """
    url = "https://www.freelancer.com/api/projects/0.1/projects/active/"
    params = {
        "query":            keyword,
        "job_details":      "true",
        "full_description": "true",
        "offset":           0,
        "limit":            FREELANCER_LIMIT,
        "sort_field":       "time_updated",
    }
    headers = {**HEADERS, "Accept": "application/json", "Freelancer-OAuth-V1": ""}

    resp = get(client, url, params=params, headers=headers)
    if not resp:
        return []

    try:
        data = resp.json()
    except Exception:
        print("    [Freelancer] Failed to parse JSON response")
        return []

    projects = data.get("result", {}).get("projects", [])
    results: list[ProjectListing] = []

    for p in projects:
        budget   = p.get("budget", {}) or {}
        currency = (p.get("currency") or {}).get("code")
        skills   = [j.get("name", "") for j in (p.get("jobs") or [])]
        seo_url  = p.get("seo_url", "")
        bid_count = (p.get("bid_stats") or {}).get("bid_count")

        # Convert unix timestamp → readable date
        ts = p.get("time_submitted")
        posted = datetime.utcfromtimestamp(ts).strftime("%Y-%m-%d") if ts else None

        results.append(ProjectListing(
            title       = p.get("title", "Untitled").strip(),
            platform    = "Freelancer.com",
            description = (p.get("description") or "")[:500].strip(),
            budget_min  = budget.get("minimum"),
            budget_max  = budget.get("maximum"),
            currency    = currency,
            skills      = [s for s in skills if s],
            bid_count   = bid_count,
            posted_date = posted,
            url         = f"https://www.freelancer.com/projects/{seo_url}" if seo_url else "https://www.freelancer.com",
            source_type = "freelance",
            keyword     = keyword,
        ))

    return results


# ─── SCRAPER: PEOPLEPERHOUR ───────────────────────────────────────────────────

def scrape_peopleperhour(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    """
    Scrapes PeoplePerHour project listing pages.

    PeoplePerHour currently renders project cards as:
        li.list__item
            div.item
                h6.item__title > a.item__url
                p.item__desc
                div.card__price
                div.card__footer-left

    Uses selectolax for fast HTML parsing.
    """

    url = f"{PPH_BASE_URL}/freelance-jobs"
    params = {
        "keyword": keyword,
        "type": "projects",
    }

    resp = get(client, url, params=params, headers=HEADERS)

    if not resp:
        return []

    # Save HTML for debugging
    safe_keyword = keyword.replace(" ", "_").replace("/", "_")
    with open(
        f"pph_debug_{safe_keyword}.html",
        "w",
        encoding="utf-8"
    ) as f:
        f.write(resp.text)

    tree = HTMLParser(resp.text)

    results: list[ProjectListing] = []

    # ---------------------------------------------------------
    # PeoplePerHour project cards
    # ---------------------------------------------------------

    cards = tree.css("li.list__item")

    print(f"    [PPH] Found {len(cards)} project cards")

    for card in cards:

        # -----------------------------------------------------
        # TITLE + URL
        # -----------------------------------------------------

        title_node = card.css_first("h6.item__title a.item__url")

        if not title_node:
            continue

        title = title_node.text(strip=True)

        href = title_node.attributes.get("href", "")

        if not href:
            continue

        if href.startswith("http"):
            link = href
        else:
            link = f"{PPH_BASE_URL}{href}"

        # -----------------------------------------------------
        # DESCRIPTION
        # -----------------------------------------------------

        desc_node = card.css_first("p.item__desc")

        description = (
            desc_node.text(strip=True)[:400]
            if desc_node
            else ""
        )

        # -----------------------------------------------------
        # BUDGET
        # -----------------------------------------------------

        budget_node = card.css_first("div.card__price")

        budget_text = (
            budget_node.text(strip=True)
            if budget_node
            else ""
        )

        budget_min, budget_max, currency = parse_budget_string(
            budget_text
        )

        # -----------------------------------------------------
        # POSTED DATE / PROPOSALS / LOCATION
        # -----------------------------------------------------

        footer = card.css_first("div.card__footer-left")

        posted = None
        bid_count = None

        if footer:

            footer_spans = footer.css("span")

            for span in footer_spans:

                text = span.text(strip=True)

                if not text:
                    continue

                # First span is normally the posted time
                if posted is None:
                    posted = text
                    continue

                # Look for proposal count
                if "proposal" in text.lower():

                    match = re.search(r"(\d+)", text)

                    if match:
                        bid_count = int(match.group(1))

        # -----------------------------------------------------
        # SKILLS
        # -----------------------------------------------------

        skills = []

        # Try common tag/skill selectors
        for selector in [
            ".skill-tag",
            ".job-tag",
            ".tag",
            ".etiquettes span",
        ]:
            for node in card.css(selector):
                text = node.text(strip=True)

                if text and text not in skills:
                    skills.append(text)

        # -----------------------------------------------------
        # CREATE LISTING
        # -----------------------------------------------------

        results.append(
            ProjectListing(
                title=title,
                platform="PeoplePerHour",
                description=description,
                budget_min=budget_min,
                budget_max=budget_max,
                currency=currency,
                skills=skills,
                bid_count=bid_count,
                posted_date=posted,
                url=link,
                source_type="freelance",
                keyword=keyword,
            )
        )

    return results

# ─── SCRAPER: TOPTAL ──────────────────────────────────────────────────────────

def scrape_toptal(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    """
    Scrapes Toptal's public job listings page.
    Toptal mostly lists longer-term engagements and full-time remote roles
    from vetted companies — good signal for B2B contract opportunities.
    """
    # Toptal's jobs page with search query
    url = f"{TOPTAL_BASE_URL}/jobs"
    params = {"q": keyword}

    resp = get(client, url, params=params, headers=HEADERS)
    if not resp:
        return []

    tree = HTMLParser(resp.text)
    results: list[ProjectListing] = []

    # Toptal job cards
    cards = (
        tree.css("li.job-listing") or
        tree.css("[data-test='job-card']") or
        tree.css(".job-item") or
        tree.css("article")
    )

    for card in cards:
        title_node = (
            card.css_first("h2 a") or
            card.css_first("h3 a") or
            card.css_first(".job-title a") or
            card.css_first("a")
        )
        if not title_node:
            continue

        title = title_node.text(strip=True)
        if not title:
            continue

        href = title_node.attrs.get("href", "")
        link = href if href.startswith("http") else f"{TOPTAL_BASE_URL}{href}"

        desc_node = (
            card.css_first(".job-description") or
            card.css_first("p") or
            card.css_first(".description")
        )
        description = desc_node.text(strip=True)[:400] if desc_node else ""

        # Toptal typically shows compensation as hourly rate
        comp_node = (
            card.css_first(".compensation") or
            card.css_first(".salary") or
            card.css_first("[data-test='compensation']")
        )
        comp_text = comp_node.text(strip=True) if comp_node else ""
        budget_min, budget_max, currency = parse_budget_string(comp_text)

        skill_nodes = card.css(".skill") or card.css(".tag") or card.css(".badge")
        skills = [s.text(strip=True) for s in skill_nodes if s.text(strip=True)]

        date_node = card.css_first("time") or card.css_first(".date")
        posted = date_node.attrs.get("datetime") or (date_node.text(strip=True) if date_node else None)

        results.append(ProjectListing(
            title       = title,
            platform    = "Toptal",
            description = description,
            budget_min  = budget_min,
            budget_max  = budget_max,
            currency    = currency or "USD",
            skills      = skills,
            bid_count   = None,
            posted_date = posted,
            url         = link,
            source_type = "freelance",
            keyword     = keyword,
        ))

    return results


# ─── SCRAPER: REMOTIVE ────────────────────────────────────────────────────────

def scrape_remotive(client: httpx.Client, keyword: str) -> list[ProjectListing]:
    """
    Calls Remotive's official, free, public REST API.
    Searches across AI-relevant categories only.
    The scoring filter in score_listings handles final relevance gating.
    """
    AI_CATEGORIES = ["machine-learning", "data", "software-dev"]
    seen_ids: set[str] = set()
    results: list[ProjectListing] = []

    for category in AI_CATEGORIES:
        params = {
            "search":   keyword,
            "category": category,
        }
        resp = get(client, REMOTIVE_API_URL, params=params, headers=HEADERS)
        if not resp:
            continue

        try:
            data = resp.json()
        except Exception:
            print("    [Remotive] Failed to parse JSON response")
            continue

        for job in data.get("jobs", []):
            job_id = str(job.get("id", ""))
            if job_id in seen_ids:
                continue
            seen_ids.add(job_id)

            salary_str = job.get("salary", "") or ""
            budget_min, budget_max, currency = parse_budget_string(salary_str)

            tags = job.get("tags") or []
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(",")]

            desc_raw = job.get("description", "") or ""
            desc_tree = HTMLParser(desc_raw)
            description = desc_tree.text(strip=True)[:500] if desc_raw else ""

            results.append(ProjectListing(
                title       = (job.get("title") or "").strip(),
                platform    = "Remotive",
                description = description,
                budget_min  = budget_min,
                budget_max  = budget_max,
                currency    = currency or "USD",
                skills      = tags,
                bid_count   = None,
                posted_date = job.get("publication_date"),
                url         = job.get("url") or "https://remotive.com",
                source_type = "remote_contract",
                keyword     = keyword,
            ))

    return results


# ─── OUTPUT ───────────────────────────────────────────────────────────────────

def print_listing(i: int, p: ProjectListing) -> None:
    """Pretty-print a single project listing to the terminal."""
    sep = "=" * 65
    print(f"\n{sep}")
    print(f"  #{i}  {p.title}")
    print(f"{sep}")
    print(f"  Platform:    {p.platform}  [{p.source_type}]")

    # Budget
    if p.budget_min and p.budget_max:
        curr = p.currency or ""
        if p.budget_min == p.budget_max:
            budget_str = f"{curr} {p.budget_min:,.0f}"
        else:
            budget_str = f"{curr} {p.budget_min:,.0f} – {p.budget_max:,.0f}"
    else:
        budget_str = "Not specified"
    print(f"  Budget:      {budget_str}")

    # Competition signal (Freelancer only)
    if p.bid_count is not None:
        signal = "🟢 Low competition" if p.bid_count < 10 else ("🟡 Medium" if p.bid_count < 25 else "🔴 High competition")
        print(f"  Bids so far: {p.bid_count}  {signal}")

    # Skills
    if p.skills:
        skills_str = ", ".join(p.skills[:7])
        if len(p.skills) > 7:
            skills_str += f" (+{len(p.skills) - 7} more)"
        print(f"  Skills:      {skills_str}")

    # Description
    if p.description:
        desc = p.description[:280]
        if len(p.description) > 280:
            desc += "..."
        print(f"  Description: {desc}")

    print(f"  Posted:      {p.posted_date or 'Unknown'}")
    print(f"  Link:        {p.url}")


def print_results(listings: list[ProjectListing]) -> None:
    """Print all listings with a summary header."""
    if not listings:
        print("\n  No results found. Try a broader query.")
        return

    # Summary by platform
    from collections import Counter
    counts = Counter(p.platform for p in listings)
    print("\n" + "─" * 65)
    print(f"  Found {len(listings)} unique projects:")
    for platform, count in counts.most_common():
        print(f"    • {platform}: {count}")
    print("─" * 65)

    for i, listing in enumerate(listings, 1):
        print_listing(i, listing)

    print(f"\n{'─' * 65}")
    print(f"  Total: {len(listings)} projects | Query ran at {datetime.now().strftime('%H:%M:%S')}")
    print("─" * 65)


ORDINALS = [
    "first", "second", "third", "fourth", "fifth",
    "sixth", "seventh", "eighth", "ninth", "tenth",
    "eleventh", "twelfth", "thirteenth", "fourteenth", "fifteenth",
    "sixteenth", "seventeenth", "eighteenth", "nineteenth", "twentieth",
]

def next_filename(results_dir: str) -> str:
    """
    Returns the next ordinal filename inside results_dir.
    Scans existing files to find the current count, then returns:
      first.json, second.json, third.json, ... twentieth.json,
      run_21.json, run_22.json, ... (fallback after 20)
    """
    os.makedirs(results_dir, exist_ok=True)
    existing = [
        f for f in os.listdir(results_dir)
        if f.endswith(".json")
    ]
    n = len(existing)  # 0-based index for next file
    if n < len(ORDINALS):
        return os.path.join(results_dir, f"{ORDINALS[n]}.json")
    return os.path.join(results_dir, f"run_{n + 1}.json")


def save_results(listings: list[ProjectListing], query: str) -> str:
    """
    Serialise results to the next ordinal JSON file in RESULTS_DIR.
    Returns the path it was saved to.
    """
    path = next_filename(RESULTS_DIR)
    data = {
        "scraped_at":    datetime.utcnow().isoformat(),
        "query":         query,
        "total_results": len(listings),
        "results":       [asdict(p) for p in listings],
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"\n  ✓ Saved {len(listings)} results → {path}")
    return path


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def next_excel_filename(excel_dir: str) -> str:
    """
    Returns the next ordinal XLSX filename inside excel_dir.
    Mirrors the JSON naming pattern: first.xlsx, second.xlsx, third.xlsx.
    """
    os.makedirs(excel_dir, exist_ok=True)
    existing = [
        f for f in os.listdir(excel_dir)
        if f.endswith(".xlsx")
    ]
    n = len(existing)
    if n < len(ORDINALS):
        return os.path.join(excel_dir, f"{ORDINALS[n]}.xlsx")
    return os.path.join(excel_dir, f"run_{n + 1}.xlsx")


def save_excel_results(listings: list[ProjectListing]) -> str | None:
    """
    Save results to the next ordinal Excel file in EXCEL_DIR.
    Returns the path it was saved to, or None if openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n  [Excel skipped] Install openpyxl to enable .xlsx export:")
        print("    pip install openpyxl")
        return None

    path = next_excel_filename(EXCEL_DIR)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Project Leads"

    headers = [
        "Title",
        "Platform",
        "Description",
        "Budget Min",
        "Budget Max",
        "Currency",
        "Skills",
        "Bid Count",
        "Posted Date",
        "URL",
        "Source Type",
        "Keyword",
        "Relevance Score",
        "Preference",
    ]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for listing in listings:
        worksheet.append([
            listing.title,
            listing.platform,
            listing.description,
            listing.budget_min,
            listing.budget_max,
            listing.currency,
            ", ".join(listing.skills),
            listing.bid_count,
            listing.posted_date,
            listing.url,
            listing.source_type,
            listing.keyword,
            listing.relevance_score,
            listing.preference_rank,
        ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    workbook.save(path)
    print(f"  Saved {len(listings)} Excel results -> {path}")
    return path


def run(query: str) -> list[ProjectListing]:
    """
    Orchestrate all scrapers for a given query.
    Always saves results to the next ordinal file in the results/ folder.
    Returns the full sorted, deduplicated list of ProjectListings.
    """
    keywords = extract_keywords(query)
    keywords = generate_project_queries(keywords)

    print(f"\n  Query:    \"{query}\"")
    print(f"  Keywords: {', '.join(keywords[:5])}" + (f" (+{len(keywords)-5} more)" if len(keywords) > 5 else ""))
    print(f"  Sources:  Freelancer.com, PeoplePerHour, Toptal, Remotive\n")

    all_listings: list[ProjectListing] = []

    # Use a single shared HTTP client for connection pooling
    with httpx.Client(follow_redirects=True) as client:

        # ── Freelancer.com ──────────────────────────────────────────────────
        print("  [1/4] Freelancer.com ...")
        for kw in keywords[:4]:  # cap at 4 keywords to avoid hammering
            results = scrape_freelancer(client, kw)
            print(f"        → '{kw}': {len(results)} projects")
            all_listings.extend(results)
            time.sleep(REQUEST_DELAY)

        # ── PeoplePerHour ───────────────────────────────────────────────────
        print("\n  [2/4] PeoplePerHour ...")
        for kw in keywords[:3]:
            results = scrape_peopleperhour(client, kw)
            print(f"        → '{kw}': {len(results)} projects")
            all_listings.extend(results)
            time.sleep(REQUEST_DELAY)

        # ── Toptal ──────────────────────────────────────────────────────────
        print("\n  [3/4] Toptal ...")
        for kw in keywords[:3]:
            results = scrape_toptal(client, kw)
            print(f"        → '{kw}': {len(results)} projects")
            all_listings.extend(results)
            time.sleep(REQUEST_DELAY)

        # ── Remotive ────────────────────────────────────────────────────────
        print("\n  [4/4] Remotive ...")
        for kw in keywords[:4]:
            results = scrape_remotive(client, kw)
            print(f"        → '{kw}': {len(results)} projects")
            all_listings.extend(results)
            time.sleep(REQUEST_DELAY)

    # Post-processing
    unique = deduplicate(all_listings)

    print(f"\n  Pre-filter: {len(unique)} unique listings")

    scored = score_listings(unique, query)

    rank_preferences(scored, top_n=5)
    sorted_ = sort_listings(scored)

    print(f"  Scored: {len(sorted_)} listings")

    print_results(sorted_)
    save_results(sorted_, query)
    save_excel_results(sorted_)
    return sorted_


# ─── ENTRY POINT ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Scrape AI project leads from Freelancer, PeoplePerHour, Toptal, and Remotive.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python main.py
  python main.py --query "computer vision object detection"
  python main.py --query "NLP chatbot LLM"
  python main.py --query "machine learning pipeline"

Results are always saved automatically to the results/ folder as:
  first.json, second.json, third.json, ...
        """
    )
    parser.add_argument(
        "--query", "-q",
        type=str,
        default=None,
        help="Search query (e.g. 'agentic computer vision projects'). Prompted if not provided.",
    )
    args = parser.parse_args()

    query = args.query
    if not query:
        print("\n  AI Project Leads Scraper")
        print("  ─────────────────────────────────────────────────────────────")
        print("  Examples: 'computer vision', 'NLP chatbot', 'ML pipeline'")
        print("  Results are auto-saved to results/first.json, second.json, ...")
        query = input("\n  Enter your search query: ").strip()
        if not query:
            print("  No query entered. Exiting.")
            raise SystemExit(0)

    run(query=query)
